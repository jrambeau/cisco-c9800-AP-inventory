import logging
from c9800 import C9800
import argparse
import pandas as pd
import datetime
import getpass
import openpyxl
import os

## Set logging level and format
logging.basicConfig(level=logging.INFO, format="%(asctime)-15s (%(levelname)s) %(message)s")

## Parse command line arguments.
# parser = argparse.ArgumentParser(description="Utility to build .xls and .csv AP inventory from C9800")
# parser.add_argument('-user', help='c9800 username', required=True)
# parser.add_argument('-password', help='c9800 password', required=True)
# parser.add_argument('-wlc_ip', help='c9800 IP address', required=True)
# args=parser.parse_args()

# Liste des IP des contrôleurs
iplist = open('devices_ip_list.txt','r').read().splitlines()

# Login
un = input('Username: ')
pw = getpass.getpass()

devices = [] # Empty array to store controllers
all_ap_dataframes = [] # List to store all AP dataframes for concatenation
all_client_dataframes = [] # List to store all client dataframes for concatenation

for ip in iplist:
    # Create an object of type C9800 and store it in the wlc variable
    devices.append(C9800(ip, un, pw))

for wlc in devices:
    # Get hostname for export files
    wlc.get_hostname()

    logging.info(f"Dealing with controller {wlc.controller_hostname}")
    timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")

    # --- AP Inventory ---
    aps = wlc.get_joined_aps()

    if aps!=0: # Check if the controller actually has APs joined. If not, we do not create the inventory files.
        # Convert AP dictionnary into dataframe
        df = pd.DataFrame(aps).T

        # Write files
        dir = "inventory"
        subdir = wlc.controller_hostname
        filename = wlc.controller_hostname + "_" + timestamp + "_AP-Inventory"
        full_path_file = os.path.join(dir, subdir, filename)
        os.makedirs(os.path.dirname(full_path_file), exist_ok=True)

        df.to_excel(full_path_file+".xlsx", index=False)
        df.to_csv(full_path_file+".csv", index=False)

        # Store controller IP alongside the dataframe for cli columns
        df["_controller_ip"] = wlc.controller_ip
        all_ap_dataframes.append(df)

    # --- Client Inventory ---
    clients = wlc.get_connected_clients()

    if clients!=0:
        df_clients = pd.DataFrame(clients).T

        dir = "inventory"
        subdir = wlc.controller_hostname
        filename = wlc.controller_hostname + "_" + timestamp + "_Client-Inventory"
        full_path_file = os.path.join(dir, subdir, filename)
        os.makedirs(os.path.dirname(full_path_file), exist_ok=True)

        df_clients.to_excel(full_path_file+".xlsx", index=False)
        df_clients.to_csv(full_path_file+".csv", index=False)

        all_client_dataframes.append(df_clients)

    # Recupération de la configuration en API. NON FONCTIONNEL, utiliser le script netmiko
    """
    # Get running-configuration
    config = wlc.get_configuration()

    # Save running-configuration to file
    dir = "configuration"
    subdir = wlc.controller_hostname
    filename = wlc.controller_hostname + "_" + datetime.datetime.now().strftime("%Y%m%d-%H%M%S") + ".conf"
    full_path_file = os.path.join(dir, subdir, filename)
    os.makedirs(os.path.dirname(full_path_file), exist_ok=True)
    with open(full_path_file, "w") as file:
        file.write(config)
    """

# Export consolidated inventory for all controllers
consolidated_timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
os.makedirs("inventory", exist_ok=True)

if all_ap_dataframes:
    df_all_aps = pd.concat(all_ap_dataframes, ignore_index=True)

    # Add CLI helper columns
    df_all_aps["cli-wlc-primary"] = df_all_aps["Controller"] + " " + df_all_aps["_controller_ip"]
    df_all_aps["cli-wlc-secondary"] = ""
    df_all_aps["cli-wlc-tertiary"] = ""
    df_all_aps["cli-ha-ap-commands"] = ""  # placeholder for CSV; formula written in xlsx below
    df_all_aps.drop(columns=["_controller_ip"], inplace=True)

    all_ap_path = os.path.join("inventory", "all-controllers-" + consolidated_timestamp + "_AP-Inventory")
    df_all_aps.to_csv(all_ap_path + ".csv", index=False)

    # Write xlsx with Excel formula in cli-ha-ap-commands
    df_all_aps.to_excel(all_ap_path + ".xlsx", index=False)
    wb = openpyxl.load_workbook(all_ap_path + ".xlsx")
    ws = wb.active
    # Find column letters for formula references
    header_row = [cell.value for cell in ws[1]]
    col_formula = header_row.index("cli-ha-ap-commands") + 1
    col_hostname_letter = openpyxl.utils.get_column_letter(header_row.index("Hostname") + 1)
    col_primary_letter = openpyxl.utils.get_column_letter(header_row.index("cli-wlc-primary") + 1)
    col_secondary_letter = openpyxl.utils.get_column_letter(header_row.index("cli-wlc-secondary") + 1)
    col_tertiary_letter = openpyxl.utils.get_column_letter(header_row.index("cli-wlc-tertiary") + 1)
    for row in range(2, ws.max_row + 1):
        r = str(row)
        h = col_hostname_letter + r
        p = col_primary_letter + r
        s = col_secondary_letter + r
        t = col_tertiary_letter + r
        formula = (
            f'="ap name "&{h}&" controller primary dummy1 1.2.3.4"&CHAR(10)&'
            f'"ap name "&{h}&" controller secondary dummy2 1.2.3.5"&CHAR(10)&'
            f'"ap name "&{h}&" controller tertiary dummy3 1.2.4.6"&CHAR(10)&'
            f'"ap name "&{h}&" controller primary "&{p}&CHAR(10)&'
            f'"ap name "&{h}&" controller secondary "&{s}&CHAR(10)&'
            f'"ap name "&{h}&" controller tertiary "&{t}&CHAR(10)'
        )
        ws.cell(row=row, column=col_formula, value=formula)
    wb.save(all_ap_path + ".xlsx")

    logging.info(f"Consolidated AP inventory exported: {all_ap_path}.csv and .xlsx ({len(df_all_aps)} APs total)")

if all_client_dataframes:
    df_all_clients = pd.concat(all_client_dataframes, ignore_index=True)
    all_client_path = os.path.join("inventory", "all-controllers-" + consolidated_timestamp + "_Client-Inventory")
    df_all_clients.to_csv(all_client_path + ".csv", index=False)
    df_all_clients.to_excel(all_client_path + ".xlsx", index=False)
    logging.info(f"Consolidated client inventory exported: {all_client_path}.csv and .xlsx ({len(df_all_clients)} clients total)")

logging.info(f"Done")
